from os.path import exists
import json
import csv
import random
import openpyxl


#EXEL
# s_list = [
#         'daveta',  'dontlivr', 'DSTroj', 'eevgenygmir', 'GrigoriyBaranchuk',
#         'Karinastriker', 'maminadochka', 'Muskulan', 'neverloaded96',
#         'psenkevich', 'quackee', 'TonyMahouney', 'VitaliaKulenko'
#     ]
# csv_data = []
# for i, student in enumerate(s_list):
#     csv_data.append((i+1, student, random.choice(s_list)))
#
# wb = openpyxl.Workbook()
# sheet = wb.active
#-----------------------------
#ручной ввод:
# c1 = sheet.cell(row=1, column=1)
# c1.value = 'XRU'
#
# c2= sheet.cell(row=1, column=2)
# c2.value = 'MUU'
#-----------------------------

# for stud_data in csv_data:
#     for col, cell in enumerate(stud_data):
#         data_cell = sheet.cell(row=stud_data[0], column=col+1)
#         data_cell.value = cell
#
# wb.save("students.xlsx")

#CSV:
# s_list = [
#         'daveta',  'dontlivr', 'DSTroj', 'eevgenygmir', 'GrigoriyBaranchuk',
#         'Karinastriker', 'maminadochka', 'Muskulan', 'neverloaded96',
#         'psenkevich', 'quackee', 'TonyMahouney', 'VitaliaKulenko'
#     ]
# csv_data = []
# for i, student in enumerate(s_list):
#     csv_data.append((i+1, student, random.choice(s_list)))
#
# print(csv_data)

# with open('students.csv', 'w') as csv_file:
#     wr = csv.writer(csv_file)
#     for stud_inf in csv_data:
#         wr.writerow(stud_inf)
#
# with open('students.csv', 'r') as csv_file:
#     rr = csv.reader(csv_file)
#     for stud_info in csv_data:
#         print(stud_info)


#JSON:
# file = "hard_work.py"
# if exists(file):
#     print("File exist")
#     students_list = []
#     with open(file) as file_name:
#         students_list = json.loads(file_name.read())
# else:
#     students_list =[
#         'daveta', 'dontlivr', 'DSTroj', 'eevgenygmir', 'GrigoriyBaranchuk',
#         'Karinastriker', 'maminadochka', 'Muskulan', 'neverloaded96',
#         'psenkevich', 'quackee', 'TonyMahouney', 'VitaliaKulenko'
#     ]
#     with open(file, "w") as file_name:
#         file_name.write(json.dumps(students_list))
#
# print(students_list)


#WORK WITH FILES
# f = open("hometask1/HT1.py", "a+")
# f.write("#GYGY\n")
# f.seek(0)  #возвращение к началу файла
# print(f.read())
# f.close()
#
# #менеджер контекста
# with open("class work5.py", "a") as file:
#     file.write('\n#print("hello, world!")')
#
# with open("class work5.py", "r") as file:
#     print(file.read())
#


